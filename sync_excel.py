import os
import json
import io
import requests
import openpyxl
import gspread
from datetime import datetime, date, time
from decimal import Decimal
from google.oauth2.service_account import Credentials

#KONFIGURASI
DROPBOX_URL_PLANTING = "https://www.dropbox.com/scl/fi/yw173h2zgnm6h3965i2ra/Database-for-Planting.xlsx?rlkey=9tfv2u3u2f12w641iauyoqc6w&dl=1"
DROPBOX_URL_FOLLOWUP = "https://www.dropbox.com/scl/fi/av13w2gl8etjfs6ai6e0i/Potential-Farmer-for-ARR.xlsx?rlkey=qeg53qzba72yguf2nt2d2ek1t&dl=1"
DROPBOX_URL_CT_COMPILE = "https://www.dropbox.com/scl/fi/bg82woa41npj9sfc6l58h/CT-Compile.xlsx?rlkey=yu29wz7rjpgvzh7vttahexc0n&dl=1"
DROPBOX_URL_BT_COMPILE = "https://www.dropbox.com/scl/fi/5au2sa502w64h32in2azi/BT-Compile.xlsx?rlkey=8do0u62sevgy3vcovrvd4ptox&dl=1"
DROPBOX_URL_COMDEV = "https://www.dropbox.com/scl/fi/bkkqvxk9nnqkxr84qry5f/ComDev-Database_Update_20260818.xlsx?rlkey=t5691lpd5x0pax5bazvf9n8st&dl=1"

TARGET_SPREADSHEET_ID = "1-LHCIUy1iD154WczY4Da_-iRaILtvL0mOW4gRW6YQac"
CAMERATRAP_SPREADSHEET_ID = "1Zkm82R3ct1HY1ov_HL35M9qEJudKPCkf0wGSNGxwdQ4"
BT_SPREADSHEET_ID = "1NJPa0nnULKqlihlkZo7J9M6gALi71LZ2_cs-0qz8SuI"
COMDEV_SPREADSHEET_ID = "1GB64g8bGV_IFjXcpRL7bu4lj9FtgRLKcUXqpNMzg50c"

SUMMARY_HEADER = [
    "No", "Grid", "KTH", "Luas", "Tipe", "Petani", "Jenis Kelamin", "Desa",
    "ID_Petak", "Tgl_Partisipasi", "Tgl_Realisasi",
    "Agro_Reg", "Agro_BPO", "Temb_Reg", "Temb_MTL",
    "Total_Tanaman_Karbon", "link",
    "Insentif_T1", "Insentif_T2", "Insentif_T3", "Insentif_T4", "Insentif_T5",
    "Insentif_T6", "Insentif_T7", "Insentif_T8", "Insentif_T9", "Insentif_T10", "Insentif_T11",
    "Nominal_T1", "Nominal_T2", "Nominal_T3", "Nominal_T4", "Nominal_T5",
    "Nominal_T6", "Nominal_T7", "Nominal_T8", "Nominal_T9", "Nominal_T10", "Nominal_T11",
    "PupukPadat", "PupukCair", "tanggal_bpo", "insentifMTL", "tanggal_mtl",
    "dusun", "luas_agro", "luas_temb", "ENL", "ENP",
]
ARR_SUMMARY_DATE_COLS = ["J", "K", "AR"]  # Tgl_Partisipasi, Tgl_Realisasi, tanggal_mtl

TARGET_CAPAIAN_MAPPING = {
    "Target Penanaman Tembawang (Ha)": "tembawang",
    "Target Penanaman Agroforestry (Ha)": "agroforestri",
    "Target Pengkayaan LOA": "LOA",
    "Target Pengkayaan PA": "PA",
}

FOLLOW_UP_HEADER = [
    "No", "Nama Petani", "Desa", "Dusun", "Lokasi", "Minat Program",
    "est_tot", "est_agr_reg", "est_tbw", "est_bpo", "est_mtl",
    "ukur_agr_reg", "ukur_tbw", "ukur_bpo", "ukur_mtl",
    "pot_agr_reg", "pot_tbw", "pot_bpo", "pot_mtl",
    "tgl_update", "status", "keterangan",
]
FOLLOW_UP_DATE_COLS = ["T"]  # tgl_update

def clean_cell(value):
    """Konversi tipe data yang tidak bisa langsung dikirim ke Google Sheets API."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, Decimal):
        return float(value)
    return value

def clean_rows(rows):
    return [[clean_cell(c) for c in row] for row in rows]

def safe_get(row, idx):
    if idx < 0 or idx >= len(row):
        return ""
    return row[idx]

def write_to_sheet(target_spreadsheet, sheet_name, data, date_cols=None):
    try:
        sheet = target_spreadsheet.worksheet(sheet_name)
    except gspread.exceptions.WorksheetNotFound:
        sheet = target_spreadsheet.add_worksheet(title=sheet_name, rows=100, cols=20)

    sheet.clear()
    if data:
        sheet.update(range_name="A1", values=data, value_input_option="USER_ENTERED")

    if date_cols and len(data) >= 2:
        for col in date_cols:
            try:
                sheet.format(f"{col}2:{col}{len(data)}", {
                    "numberFormat": {"type": "DATE", "pattern": "m/d/yyyy"}
                })
            except Exception:
                pass

    return sheet

def download_workbook(url, label):
    response = requests.get(url)
    response.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)

def process_arr(source_workbook, target_spreadsheet):
    if "ARR" not in source_workbook.sheetnames:
        return

    raw_data = [list(row) for row in source_workbook["ARR"].iter_rows(values_only=True)]

    header_row_idx = next(
        (i for i, row in enumerate(raw_data)
         if all(k in [str(v).strip() if v is not None else "" for v in row]
                for k in ("No", "Kompartemen", "Nama KTH"))),
        None,
    )
    if header_row_idx is None:
        raise ValueError("Header tidak ditemukan di sheet ARR")

    data = raw_data[header_row_idx:]
    header = [str(v).strip() if v is not None else "" for v in data[0]]
    if "No" not in header:
        raise ValueError("Kolom 'No' tidak ditemukan di sheet ARR")
    no_col_index = header.index("No")

    data = [row[no_col_index:] for row in data]
    data = [row for row in data if any(str(c).strip() != "" for c in row if c is not None)]
    data = [["" if c == "#REF!" else c for c in row] for row in data]

    if len(data) >= 2:
        del data[1]  # baris kedua adalah baris kosong/sub-header sisa

    write_to_sheet(target_spreadsheet, "ARR", clean_rows(data))

    summary = [SUMMARY_HEADER]
    for row in data[1:]:
        gender = "Laki-laki" if safe_get(row, 8) == 1 else "Perempuan" if safe_get(row, 9) == 1 else ""
        tipe = {"TBW": "Tembawang", "AGR": "Agroforestri"}.get(safe_get(row, 5), safe_get(row, 5))

        summary_row = [
            safe_get(row, 0), safe_get(row, 1), safe_get(row, 3), safe_get(row, 6),
            tipe, safe_get(row, 7), gender, safe_get(row, 11),
            safe_get(row, 13), safe_get(row, 14), safe_get(row, 15),
            safe_get(row, 21), safe_get(row, 22), safe_get(row, 23), safe_get(row, 24),
            safe_get(row, 48), safe_get(row, 368),
            safe_get(row, 334), safe_get(row, 337), safe_get(row, 340),
            safe_get(row, 343), safe_get(row, 346), safe_get(row, 349),
            safe_get(row, 352), safe_get(row, 355), safe_get(row, 358),
            safe_get(row, 361), safe_get(row, 364),
            safe_get(row, 336), safe_get(row, 339), safe_get(row, 342),
            safe_get(row, 345), safe_get(row, 348), safe_get(row, 351),
            safe_get(row, 354), safe_get(row, 357), safe_get(row, 360),
            safe_get(row, 363), safe_get(row, 366),
            safe_get(row, 39), safe_get(row, 40), safe_get(row, 41),
            safe_get(row, 43), safe_get(row, 44),
            safe_get(row, 12), safe_get(row, 16), safe_get(row, 17),
            safe_get(row, 18), safe_get(row, 19),
        ]
        summary.append(clean_rows([summary_row])[0])

    write_to_sheet(target_spreadsheet, "ARR_SUMMARY", summary, ARR_SUMMARY_DATE_COLS)

def process_target_capaian(source_workbook, target_spreadsheet):
    if "Target Capaian" not in source_workbook.sheetnames:
        return

    raw_data = [list(row) for row in source_workbook["Target Capaian"].iter_rows(values_only=True)]
    data = raw_data[1:]  # skip judul besar
    if not data:
        return

    for row in data[1:]:
        uraian = str(row[0]).strip() if row[0] is not None else ""
        if uraian in TARGET_CAPAIAN_MAPPING:
            row[0] = TARGET_CAPAIAN_MAPPING[uraian]

    write_to_sheet(target_spreadsheet, "TARGET_CAPAIAN", clean_rows(data))

def process_cumulative_area(source_workbook, target_spreadsheet):
    sheet_name = "Cumulative Area"
    if sheet_name not in source_workbook.sheetnames:
        return

    raw_data = [list(row) for row in source_workbook[sheet_name].iter_rows(values_only=True)]

    data = [row[1:] for row in raw_data[3:]]

    write_to_sheet(target_spreadsheet, "Planting Cumulative Area", clean_rows(data))

def process_ct_compile(source_workbook, target_spreadsheet):
    sheet_name = "CT Compile"
    if sheet_name not in source_workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' tidak ditemukan. "
            f"Sheet yang tersedia: {source_workbook.sheetnames}"
        )

    data = [list(row) for row in source_workbook[sheet_name].iter_rows(values_only=True)]
    if not data:
        raise ValueError(f"Sheet '{sheet_name}' kosong, tidak ada data untuk ditulis.")

    write_to_sheet(target_spreadsheet, "db-CT Compile", clean_rows(data))

def process_bt_compile(source_workbook, target_spreadsheet):
    sheet_name = "BT"
    if sheet_name not in source_workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' tidak ditemukan. "
            f"Sheet yang tersedia: {source_workbook.sheetnames}"
        )

    data = [list(row) for row in source_workbook[sheet_name].iter_rows(values_only=True)]
    if not data:
        raise ValueError(f"Sheet '{sheet_name}' kosong, tidak ada data untuk ditulis.")

    write_to_sheet(target_spreadsheet, "db-BT", clean_rows(data))

def process_comdev(source_workbook, target_spreadsheet):
    sheet_name = "ComDev"
    if sheet_name not in source_workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' tidak ditemukan. "
            f"Sheet yang tersedia: {source_workbook.sheetnames}"
        )

    raw_data = [list(row) for row in source_workbook[sheet_name].iter_rows(values_only=True)]

    data = [row[1:] for row in raw_data[3:]]
    if not data:
        raise ValueError(f"Sheet '{sheet_name}' kosong, tidak ada data untuk ditulis.")

    write_to_sheet(target_spreadsheet, "db-comdev", clean_rows(data))

def process_follow_up_petani(source_workbook, target_spreadsheet):
    if "Follow Up Petani" not in source_workbook.sheetnames:
        return

    data = [list(row) for row in source_workbook["Follow Up Petani"].iter_rows(values_only=True)]
    result = [FOLLOW_UP_HEADER]

    for row in data[3:]:
        if not (safe_get(row, 1) or safe_get(row, 2) or safe_get(row, 3)):
            continue
        result.append([safe_get(row, i) for i in range(1, 23)])

    write_to_sheet(target_spreadsheet, "FOLLOW_UP_PETANI", clean_rows(result), FOLLOW_UP_DATE_COLS)

def main():
    creds_dict = json.loads(os.environ["GOOGLE_CREDENTIALS"])
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    gc = gspread.authorize(creds)
    target_spreadsheet = gc.open_by_key(TARGET_SPREADSHEET_ID)

    planting_wb = download_workbook(DROPBOX_URL_PLANTING, "Database-for-Planting")
    process_arr(planting_wb, target_spreadsheet)
    process_target_capaian(planting_wb, target_spreadsheet)
    process_cumulative_area(planting_wb, target_spreadsheet)

    followup_wb = download_workbook(DROPBOX_URL_FOLLOWUP, "Potential-Farmer-for-ARR")
    process_follow_up_petani(followup_wb, target_spreadsheet)

    ct_compile_wb = download_workbook(DROPBOX_URL_CT_COMPILE, "CT Compile")
    cameratrap_spreadsheet = gc.open_by_key(CAMERATRAP_SPREADSHEET_ID)
    process_ct_compile(ct_compile_wb, cameratrap_spreadsheet)

    bt_compile_wb = download_workbook(DROPBOX_URL_BT_COMPILE, "BT Compile")
    bt_spreadsheet = gc.open_by_key(BT_SPREADSHEET_ID)
    process_bt_compile(bt_compile_wb, bt_spreadsheet)

    comdev_wb = download_workbook(DROPBOX_URL_COMDEV, "ComDev Database")
    comdev_spreadsheet = gc.open_by_key(COMDEV_SPREADSHEET_ID)
    process_comdev(comdev_wb, comdev_spreadsheet)


if __name__ == "__main__":
    main()
